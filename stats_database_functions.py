import openpyxl

statsWorkbook = openpyxl.load_workbook('stats_database.xlsx')
normalKickoffs = statsWorkbook.worksheets[0]
squibKickoffs = statsWorkbook.worksheets[1]
onsideKickoffs = statsWorkbook.worksheets[2]
punts = statsWorkbook.worksheets[3]

def updateNormalKickoffResult(result):
    """
    Update the stats sheet with the result of the normal kick

    """

    result = str(result[0])
    rowNum = 1

    # Find the latest row
    for cell in normalKickoffs['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1

    # Add result
    normalKickoffs.cell(row=rowNum, column=1).value = str(result)

    # Update number of results
    numResults = rowNum - 1
    normalKickoffs.cell(row=3, column=3).value = numResults

    # Update the number of that specific result and the percentages
    if str(result) == "Touchdown":
        prevNum = normalKickoffs.cell(row=3, column=4).value
        normalKickoffs.cell(row=3, column=4).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=4).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Fumble":
        prevNum = normalKickoffs.cell(row=3, column=5).value
        normalKickoffs.cell(row=3, column=5).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=5).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "5":
        prevNum = normalKickoffs.cell(row=3, column=6).value
        normalKickoffs.cell(row=3, column=6).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=6).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "10":
        prevNum = normalKickoffs.cell(row=3, column=7).value
        normalKickoffs.cell(row=3, column=7).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=7).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "20":
        prevNum = normalKickoffs.cell(row=3, column=8).value
        normalKickoffs.cell(row=3, column=8).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=8).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "25":
        prevNum = normalKickoffs.cell(row=3, column=9).value
        normalKickoffs.cell(row=3, column=9).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=9).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "30":
        prevNum = normalKickoffs.cell(row=3, column=10).value
        normalKickoffs.cell(row=3, column=10).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=10).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "40":
        prevNum = normalKickoffs.cell(row=3, column=11).value
        normalKickoffs.cell(row=3, column=11).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=11).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "50":
        prevNum = normalKickoffs.cell(row=3, column=12).value
        normalKickoffs.cell(row=3, column=12).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=12).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "65":
        prevNum = normalKickoffs.cell(row=3, column=13).value
        normalKickoffs.cell(row=3, column=13).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=13).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Returned TD":
        prevNum = normalKickoffs.cell(row=3, column=14).value
        normalKickoffs.cell(row=3, column=14).value = int(prevNum) + 1
        normalKickoffs.cell(row=4, column=14).value = ((int(prevNum) + 1)/numResults) * 100
    statsWorkbook.save('stats_database.xlsx')

def updateSquibKickoffResult(result):
    """
    Update the stats sheet with the result of the squib kick

    """

    result = str(result[0])
    rowNum = 1

    # Find the latest row
    for cell in squibKickoffs['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1

    # Add result
    squibKickoffs.cell(row=rowNum, column=1).value = str(result)

    # Update number of results
    numResults = rowNum - 1
    squibKickoffs.cell(row=3, column=3).value = numResults

    # Update the number of that specific result and the percentages
    if str(result) == "Touchdown":
        prevNum = squibKickoffs.cell(row=3, column=4).value
        squibKickoffs.cell(row=3, column=4).value = int(prevNum) + 1
        squibKickoffs.cell(row=4, column=4).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Fumble":
        prevNum = squibKickoffs.cell(row=3, column=5).value
        squibKickoffs.cell(row=3, column=5).value = int(prevNum) + 1
        squibKickoffs.cell(row=4, column=5).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "30":
        prevNum = squibKickoffs.cell(row=3, column=6).value
        squibKickoffs.cell(row=3, column=6).value = int(prevNum) + 1
        squibKickoffs.cell(row=4, column=6).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "35":
        prevNum = squibKickoffs.cell(row=3, column=7).value
        squibKickoffs.cell(row=3, column=7).value = int(prevNum) + 1
        squibKickoffs.cell(row=4, column=7).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "40":
        prevNum = squibKickoffs.cell(row=3, column=8).value
        squibKickoffs.cell(row=3, column=8).value = int(prevNum) + 1
        squibKickoffs.cell(row=4, column=8).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "45":
        prevNum = squibKickoffs.cell(row=3, column=9).value
        squibKickoffs.cell(row=3, column=9).value = int(prevNum) + 1
        squibKickoffs.cell(row=4, column=9).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "50":
        prevNum = squibKickoffs.cell(row=3, column=10).value
        squibKickoffs.cell(row=3, column=10).value = int(prevNum) + 1
        squibKickoffs.cell(row=4, column=10).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Returned TD":
        prevNum = squibKickoffs.cell(row=3, column=11).value
        squibKickoffs.cell(row=3, column=11).value = int(prevNum) + 1
        squibKickoffs.cell(row=4, column=11).value = ((int(prevNum) + 1)/numResults) * 100
    statsWorkbook.save('stats_database.xlsx')

def updateOnsideKickoffResult(result):
    """
    Update the stats sheet with the result of the onside kick

    """

    result = str(result[0])
    rowNum = 1

    # Find the latest row
    for cell in onsideKickoffs['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1

    # Add result
    onsideKickoffs.cell(row=rowNum, column=1).value = str(result)

    # Update number of results
    numResults = rowNum - 1
    onsideKickoffs.cell(row=3, column=3).value = numResults

    # Update the number of that specific result and the percentages
    if str(result) == "Recovered":
        prevNum = onsideKickoffs.cell(row=3, column=4).value
        onsideKickoffs.cell(row=3, column=4).value = int(prevNum) + 1
        onsideKickoffs.cell(row=4, column=4).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "No Good":
        prevNum = onsideKickoffs.cell(row=3, column=5).value
        onsideKickoffs.cell(row=3, column=5).value = int(prevNum) + 1
        onsideKickoffs.cell(row=4, column=5).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Returned TD":
        prevNum = onsideKickoffs.cell(row=3, column=6).value
        onsideKickoffs.cell(row=3, column=6).value = int(prevNum) + 1
        onsideKickoffs.cell(row=4, column=6).value = ((int(prevNum) + 1)/numResults) * 100
    statsWorkbook.save('stats_database.xlsx')

def updatePuntResult(result):
    """
    Update the stats sheet with the result of the normal kick

    """

    result = str(result[0])
    rowNum = 1

    # Find the latest row
    for cell in normalKickoffs['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1

    # Add result
    punts.cell(row=rowNum, column=1).value = str(result)

    # Update number of results
    numResults = rowNum - 1
    punts.cell(row=3, column=3).value = numResults

    # Update the number of that specific result and the percentages
    if str(result) == "Punt Six":
        prevNum = punts.cell(row=3, column=4).value
        punts.cell(row=3, column=4).value = int(prevNum) + 1
        punts.cell(row=4, column=4).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Blocked":
        prevNum = punts.cell(row=3, column=5).value
        punts.cell(row=3, column=5).value = int(prevNum) + 1
        punts.cell(row=4, column=5).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "5":
        prevNum = punts.cell(row=3, column=6).value
        punts.cell(row=3, column=6).value = int(prevNum) + 1
        punts.cell(row=4, column=6).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "10":
        prevNum = punts.cell(row=3, column=7).value
        punts.cell(row=3, column=7).value = int(prevNum) + 1
        punts.cell(row=4, column=7).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "15":
        prevNum = punts.cell(row=3, column=8).value
        punts.cell(row=3, column=8).value = int(prevNum) + 1
        punts.cell(row=4, column=8).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "20":
        prevNum = punts.cell(row=3, column=9).value
        punts.cell(row=3, column=9).value = int(prevNum) + 1
        punts.cell(row=4, column=9).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "25":
        prevNum = punts.cell(row=3, column=10).value
        punts.cell(row=3, column=10).value = int(prevNum) + 1
        punts.cell(row=4, column=10).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "30":
        prevNum = punts.cell(row=3, column=11).value
        punts.cell(row=3, column=11).value = int(prevNum) + 1
        punts.cell(row=4, column=11).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "35":
        prevNum = punts.cell(row=3, column=12).value
        punts.cell(row=3, column=12).value = int(prevNum) + 1
        punts.cell(row=4, column=12).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "40":
        prevNum = punts.cell(row=3, column=13).value
        punts.cell(row=3, column=13).value = int(prevNum) + 1
        punts.cell(row=4, column=13).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "45":
        prevNum = punts.cell(row=3, column=14).value
        punts.cell(row=3, column=14).value = int(prevNum) + 1
        punts.cell(row=4, column=14).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "50":
        prevNum = punts.cell(row=3, column=15).value
        punts.cell(row=3, column=15).value = int(prevNum) + 1
        punts.cell(row=4, column=15).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "55":
        prevNum = punts.cell(row=3, column=16).value
        punts.cell(row=3, column=16).value = int(prevNum) + 1
        punts.cell(row=4, column=16).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "60":
        prevNum = punts.cell(row=3, column=17).value
        punts.cell(row=3, column=17).value = int(prevNum) + 1
        punts.cell(row=4, column=17).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "65":
        prevNum = punts.cell(row=3, column=18).value
        punts.cell(row=3, column=18).value = int(prevNum) + 1
        punts.cell(row=4, column=18).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "70":
        prevNum = punts.cell(row=3, column=19).value
        punts.cell(row=3, column=19).value = int(prevNum) + 1
        punts.cell(row=4, column=19).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Touchback":
        prevNum = punts.cell(row=3, column=20).value
        punts.cell(row=3, column=20).value = int(prevNum) + 1
        punts.cell(row=4, column=20).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Fumble":
        prevNum = punts.cell(row=3, column=20).value
        punts.cell(row=3, column=20).value = int(prevNum) + 1
        punts.cell(row=4, column=20).value = ((int(prevNum) + 1)/numResults) * 100
    elif str(result) == "Touchdown":
        prevNum = punts.cell(row=3, column=20).value
        punts.cell(row=3, column=20).value = int(prevNum) + 1
        punts.cell(row=4, column=20).value = ((int(prevNum) + 1)/numResults) * 100
    statsWorkbook.save('stats_database.xlsx')