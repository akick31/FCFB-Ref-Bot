import openpyxl

statsWorkbook = openpyxl.load_workbook('stats_database.xlsx')
normalKickoffs = statsWorkbook.worksheets[0]
squibKickoffs = statsWorkbook.worksheets[1]
onsideKickoffs = statsWorkbook.worksheets[2]
punts = statsWorkbook.worksheets[3]

def updateNormalKickoffResult(result):
    """
    Update the stats sheet with the result of the normal kick

    """

    rowNum = 1

    # Find the latest row
    for cell in normalKickoffs['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1

    # Add result
    normalKickoffs.cell(row=rowNum, column=1).value = str(result)

    # Update number of results
    numResults = rowNum - 1
    normalKickoffs.cell(row=3, column=3).value = numResults

    # Update the number of that specific result and the percentages
    if str(result) == "Touchdown":
        prevNum = int(normalKickoffs.cell(row=3, column=4).value)
        normalKickoffs.cell(row=3, column=4).value = prevNum + 1
        normalKickoffs.cell(row=4, column=4).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Fumble":
        prevNum = int(normalKickoffs.cell(row=3, column=5).value)
        normalKickoffs.cell(row=3, column=5).value = prevNum + 1
        normalKickoffs.cell(row=4, column=5).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "5":
        prevNum = int(normalKickoffs.cell(row=3, column=6).value)
        normalKickoffs.cell(row=3, column=6).value = prevNum + 1
        normalKickoffs.cell(row=4, column=6).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "10":
        prevNum = int(normalKickoffs.cell(row=3, column=7).value)
        normalKickoffs.cell(row=3, column=7).value = prevNum + 1
        normalKickoffs.cell(row=4, column=7).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "20":
        prevNum = int(normalKickoffs.cell(row=3, column=8).value)
        normalKickoffs.cell(row=3, column=8).value = prevNum + 1
        normalKickoffs.cell(row=4, column=8).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "25":
        prevNum = int(normalKickoffs.cell(row=3, column=9).value)
        normalKickoffs.cell(row=3, column=9).value = prevNum + 1
        normalKickoffs.cell(row=4, column=9).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "30":
        prevNum = int(normalKickoffs.cell(row=3, column=10).value)
        normalKickoffs.cell(row=3, column=10).value = prevNum + 1
        normalKickoffs.cell(row=4, column=10).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "40":
        prevNum = int(normalKickoffs.cell(row=3, column=11).value)
        normalKickoffs.cell(row=3, column=11).value = prevNum + 1
        normalKickoffs.cell(row=4, column=11).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "50":
        prevNum = int(normalKickoffs.cell(row=3, column=12).value)
        normalKickoffs.cell(row=3, column=12).value = prevNum + 1
        normalKickoffs.cell(row=4, column=12).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "65":
        prevNum = int(normalKickoffs.cell(row=3, column=13).value)
        normalKickoffs.cell(row=3, column=13).value = prevNum + 1
        normalKickoffs.cell(row=4, column=13).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Returned TD":
        prevNum = int(normalKickoffs.cell(row=3, column=14).value)
        normalKickoffs.cell(row=3, column=14).value = prevNum + 1
        normalKickoffs.cell(row=4, column=14).value = ((prevNum + 1)/numResults) * 100

def updateSquibKickoffResult(result):
    """
    Update the stats sheet with the result of the squib kick

    """

    rowNum = 1

    # Find the latest row
    for cell in squibKickoffs['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1

    # Add result
    squibKickoffs.cell(row=rowNum, column=1).value = str(result)

    # Update number of results
    numResults = rowNum - 1
    squibKickoffs.cell(row=3, column=3).value = numResults

    # Update the number of that specific result and the percentages
    if str(result) == "Touchdown":
        prevNum = int(squibKickoffs.cell(row=3, column=4).value)
        squibKickoffs.cell(row=3, column=4).value = prevNum + 1
        squibKickoffs.cell(row=4, column=4).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Fumble":
        prevNum = int(squibKickoffs.cell(row=3, column=5).value)
        squibKickoffs.cell(row=3, column=5).value = prevNum + 1
        squibKickoffs.cell(row=4, column=5).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "30":
        prevNum = int(squibKickoffs.cell(row=3, column=6).value)
        squibKickoffs.cell(row=3, column=6).value = prevNum + 1
        squibKickoffs.cell(row=4, column=6).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "35":
        prevNum = int(squibKickoffs.cell(row=3, column=7).value)
        squibKickoffs.cell(row=3, column=7).value = prevNum + 1
        squibKickoffs.cell(row=4, column=7).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "40":
        prevNum = int(squibKickoffs.cell(row=3, column=8).value)
        squibKickoffs.cell(row=3, column=8).value = prevNum + 1
        squibKickoffs.cell(row=4, column=8).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "45":
        prevNum = int(squibKickoffs.cell(row=3, column=9).value)
        squibKickoffs.cell(row=3, column=9).value = prevNum + 1
        squibKickoffs.cell(row=4, column=9).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "50":
        prevNum = int(squibKickoffs.cell(row=3, column=10).value)
        squibKickoffs.cell(row=3, column=10).value = prevNum + 1
        squibKickoffs.cell(row=4, column=10).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Returned TD":
        prevNum = int(squibKickoffs.cell(row=3, column=11).value)
        squibKickoffs.cell(row=3, column=11).value = prevNum + 1
        squibKickoffs.cell(row=4, column=11).value = ((prevNum + 1)/numResults) * 100

def updateOnsideKickoffResult(result):
    """
    Update the stats sheet with the result of the onside kick

    """

    rowNum = 1

    # Find the latest row
    for cell in onsideKickoffs['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1

    # Add result
    onsideKickoffs.cell(row=rowNum, column=1).value = str(result)

    # Update number of results
    numResults = rowNum - 1
    onsideKickoffs.cell(row=3, column=3).value = numResults

    # Update the number of that specific result and the percentages
    if str(result) == "Recovered":
        prevNum = int(onsideKickoffs.cell(row=3, column=4).value)
        onsideKickoffs.cell(row=3, column=4).value = prevNum + 1
        onsideKickoffs.cell(row=4, column=4).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "No Good":
        prevNum = int(onsideKickoffs.cell(row=3, column=5).value)
        onsideKickoffs.cell(row=3, column=5).value = prevNum + 1
        onsideKickoffs.cell(row=4, column=5).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Returned TD":
        prevNum = int(onsideKickoffs.cell(row=3, column=6).value)
        onsideKickoffs.cell(row=3, column=6).value = prevNum + 1
        onsideKickoffs.cell(row=4, column=6).value = ((prevNum + 1)/numResults) * 100

def updatePuntResult(result):
    """
    Update the stats sheet with the result of the normal kick

    """

    rowNum = 1

    # Find the latest row
    for cell in normalKickoffs['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1

    # Add result
    punts.cell(row=rowNum, column=1).value = str(result)

    # Update number of results
    numResults = rowNum - 1
    punts.cell(row=3, column=3).value = numResults

    # Update the number of that specific result and the percentages
    if str(result) == "Punt Six":
        prevNum = int(punts.cell(row=3, column=4).value)
        punts.cell(row=3, column=4).value = prevNum + 1
        punts.cell(row=4, column=4).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Blocked":
        prevNum = int(punts.cell(row=3, column=5).value)
        punts.cell(row=3, column=5).value = prevNum + 1
        punts.cell(row=4, column=5).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "5":
        prevNum = int(punts.cell(row=3, column=6).value)
        punts.cell(row=3, column=6).value = prevNum + 1
        punts.cell(row=4, column=6).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "10":
        prevNum = int(punts.cell(row=3, column=7).value)
        punts.cell(row=3, column=7).value = prevNum + 1
        punts.cell(row=4, column=7).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "15":
        prevNum = int(punts.cell(row=3, column=8).value)
        punts.cell(row=3, column=8).value = prevNum + 1
        punts.cell(row=4, column=8).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "20":
        prevNum = int(punts.cell(row=3, column=9).value)
        punts.cell(row=3, column=9).value = prevNum + 1
        punts.cell(row=4, column=9).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "25":
        prevNum = int(punts.cell(row=3, column=10).value)
        punts.cell(row=3, column=10).value = prevNum + 1
        punts.cell(row=4, column=10).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "30":
        prevNum = int(punts.cell(row=3, column=11).value)
        punts.cell(row=3, column=11).value = prevNum + 1
        punts.cell(row=4, column=11).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "35":
        prevNum = int(punts.cell(row=3, column=12).value)
        punts.cell(row=3, column=12).value = prevNum + 1
        punts.cell(row=4, column=12).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "40":
        prevNum = int(punts.cell(row=3, column=13).value)
        punts.cell(row=3, column=13).value = prevNum + 1
        punts.cell(row=4, column=13).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "45":
        prevNum = int(punts.cell(row=3, column=14).value)
        punts.cell(row=3, column=14).value = prevNum + 1
        punts.cell(row=4, column=14).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "50":
        prevNum = int(punts.cell(row=3, column=15).value)
        punts.cell(row=3, column=15).value = prevNum + 1
        punts.cell(row=4, column=15).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "55":
        prevNum = int(punts.cell(row=3, column=16).value)
        punts.cell(row=3, column=16).value = prevNum + 1
        punts.cell(row=4, column=16).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "60":
        prevNum = int(punts.cell(row=3, column=17).value)
        punts.cell(row=3, column=17).value = prevNum + 1
        punts.cell(row=4, column=17).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "65":
        prevNum = int(punts.cell(row=3, column=18).value)
        punts.cell(row=3, column=18).value = prevNum + 1
        punts.cell(row=4, column=18).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "70":
        prevNum = int(punts.cell(row=3, column=19).value)
        punts.cell(row=3, column=19).value = prevNum + 1
        punts.cell(row=4, column=19).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Touchback":
        prevNum = int(punts.cell(row=3, column=20).value)
        punts.cell(row=3, column=20).value = prevNum + 1
        punts.cell(row=4, column=20).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Fumble":
        prevNum = int(punts.cell(row=3, column=20).value)
        punts.cell(row=3, column=20).value = prevNum + 1
        punts.cell(row=4, column=20).value = ((prevNum + 1)/numResults) * 100
    elif str(result) == "Touchdown":
        prevNum = int(punts.cell(row=3, column=20).value)
        punts.cell(row=3, column=20).value = prevNum + 1
        punts.cell(row=4, column=20).value = ((prevNum + 1)/numResults) * 100