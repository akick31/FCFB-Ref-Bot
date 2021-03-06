import openpyxl

userWorkbook = openpyxl.load_workbook('user_database.xlsx')
userDB = userWorkbook.worksheets[0]

"""
Functions that modify the database or retrieve from the user database

@author: apkick
"""

def getNickname(team):
    """
    Get a team's nickname
    
    """
    
    rowNum = 1
    for cell in userDB['A']:
        if cell.value.lower() == team.lower():
            return userDB.cell(row = rowNum, column = 2).value
        else:
            rowNum = rowNum + 1
    return "COULD NOT FIND"


def getUser(team):
    """
    Get a user for a team
    
    """
    
    rowNum = 1
    for cell in userDB['A']:
        if cell.value.lower() == team.lower():
            return userDB.cell(row = rowNum, column = 4).value
        else:
            rowNum = rowNum + 1
    return "COULD NOT FIND"


def getOffensivePlaybook(team):
    """
    Get a team's offensive playbook
    
    """
    
    rowNum = 1
    for cell in userDB['A']:
        if cell.value.lower() == team.lower():
            return userDB.cell(row = rowNum, column = 6).value
        else:
            rowNum = rowNum + 1
    return "COULD NOT FIND"


def getDefensivePlaybook(team):
    """
    Get a team's defensive playbook
    
    """
    
    rowNum = 1
    for cell in userDB['A']:
        if cell.value.lower() == team.lower():
            return userDB.cell(row = rowNum, column = 7).value
        else:
            rowNum = rowNum + 1
    return "COULD NOT FIND"


def getRecord(team):
    """
    Get a team's record
    
    """ 
    
    rowNum = 1
    for cell in userDB['A']:
        if cell.value.lower() == team.lower():
            return userDB.cell(row = rowNum, column = 8).value
        else:
            rowNum = rowNum + 1
    return "COULD NOT FIND"

def getTeamInformation(team):
    """
    Get the team's entire list of information from the db

    """
    rowNum = 1
    for cell in userDB['A']:
        if cell.value.lower() == team.lower():
            teamInfo = []
            for i in range(1, 8):
                teamInfo.append(userDB.cell(row = rowNum, column = i).value)
            return teamInfo
        else:
            rowNum = rowNum + 1
    return "COULD NOT FIND"


def updateRecord(team, result):
    """
    Update the team's record based on their game result

    """
    rowNum = 1
    record = None
    for cell in userDB['A']:
        if cell.value.lower() == team.lower():
            record = userDB.cell(row = rowNum, column = 8).value
        else:
            rowNum = rowNum + 1

    if record is not None:
        wins = record.split["-"][0]
        losses = record.split["-"][1]
        ties = record.split["-"][2]
        if result == "W":
            record = str(int(wins) + 1) + "-" + losses + "-" + ties
        elif result == "L":
            record = wins + "-" + str(int(losses) + 1) + "-" + ties
        elif result == "T":
            record = wins + "-" + losses + "-" + str(int(ties) + 1)
        userDB.cell(row = rowNum, column = 8).value = record
    else:
        return "COULD NOT FIND"


def deleteTeam(team):
    """
    Iterate through the teams and verify the user does not already exist.
    
    Returns True if they exist, returns False if they do not
    
    """ 
    
    rowNum = 1
    for cell in userDB['A']:
        if cell.value.lower() == team.lower():
            userDB.delete_rows(rowNum)
            userWorkbook.save('user_database.xlsx')
            return "SUCCESS"
        else:
            rowNum = rowNum + 1
    return "COULD NOT FIND"


def addUser(teamInfo):
    """
    Add a new user to the database
    
    """ 
    
    # Find the first empty row
    rowNum = 1
    for cell in userDB['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1
    
    userDB.cell(row = rowNum, column = 1).value = teamInfo[0]
    userDB.cell(row = rowNum, column = 2).value = teamInfo[1]
    userDB.cell(row = rowNum, column = 3).value = teamInfo[2]
    userDB.cell(row = rowNum, column = 4).value = teamInfo[3]
    userDB.cell(row = rowNum, column = 5).value = teamInfo[4]
    userDB.cell(row = rowNum, column = 6).value = teamInfo[5]
    userDB.cell(row = rowNum, column = 7).value = teamInfo[6]
    userDB.cell(row = rowNum, column = 8).value = "0-0"
    userWorkbook.save('user_database.xlsx')


def checkUser(user):
    """
    Iterate through the teams and verify the user does not already exist.
    
    Returns True if they exist, returns False if they do not
    
    """ 
    
    for cell in userDB['D']:
        if cell.value.lower() == user.lower():
            return True
        elif cell.value is None or cell.value == "":
            break
    
    return False


def checkName(name):
    """
    Iterate through the teams and verify the team name does not already exist.
    
    Returns True if they exist, returns False if they do not
    
    """ 
    
    for cell in userDB['A']:
        if cell.value.lower() == name.lower():
            return True
        elif cell.value is None or cell.value == "":
            break
    
    return False
