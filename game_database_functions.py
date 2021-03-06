import openpyxl

openpyxlGameWorkbook = openpyxl.load_workbook('game_database.xlsx')
ongoingGames = openpyxlGameWorkbook.worksheets[0]
finishedGames = openpyxlGameWorkbook.worksheets[1]


"""
Functions that modify the database or retrieve from the game database

@author: apkick
"""

def addGameToDatabase(channel, homeTeamInfo, awayTeamInfo):
    """
    Add the game to the database
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    ongoingGames.cell(row = rowNum, column = 1).value = str(channel.id)
    ongoingGames.cell(row = rowNum, column = 2).value = homeTeamInfo["name"]
    ongoingGames.cell(row = rowNum, column = 3).value = homeTeamInfo["nickname"]
    ongoingGames.cell(row = rowNum, column = 4).value = homeTeamInfo["user"]
    ongoingGames.cell(row = rowNum, column = 5).value = homeTeamInfo["offensive playbook"]
    ongoingGames.cell(row = rowNum, column = 6).value = homeTeamInfo["defensive playbook"]
    ongoingGames.cell(row = rowNum, column = 7).value = 0  # home score
    ongoingGames.cell(row = rowNum, column = 8).value = 3  # home timeouts
    ongoingGames.cell(row = rowNum, column = 9).value = 0  # home total yards
    ongoingGames.cell(row = rowNum, column = 10).value = 0  # home passing yards
    ongoingGames.cell(row = rowNum, column = 11).value = 0  # home rushing yards
    ongoingGames.cell(row = rowNum, column = 12).value = 0  # home interceptions
    ongoingGames.cell(row = rowNum, column = 13).value = 0  # home fumbles
    ongoingGames.cell(row = rowNum, column = 14).value = awayTeamInfo["name"]
    ongoingGames.cell(row = rowNum, column = 15).value = awayTeamInfo["nickname"]
    ongoingGames.cell(row = rowNum, column = 16).value = awayTeamInfo["user"]
    ongoingGames.cell(row = rowNum, column = 17).value = awayTeamInfo["offensive playbook"]
    ongoingGames.cell(row = rowNum, column = 18).value = awayTeamInfo["defensive playbook"]
    ongoingGames.cell(row = rowNum, column = 19).value = 0  # away score
    ongoingGames.cell(row = rowNum, column = 20).value = 3  # away timeouts
    ongoingGames.cell(row = rowNum, column = 21).value = 0  # away total yards
    ongoingGames.cell(row = rowNum, column = 22).value = 0  # away passing yards
    ongoingGames.cell(row = rowNum, column = 23).value = 0  # away rushing yards
    ongoingGames.cell(row = rowNum, column = 24).value = 0  # away interceptions
    ongoingGames.cell(row = rowNum, column = 25).value = 0  # away fumbles
    ongoingGames.cell(row = rowNum, column = 26).value = "NONE"  # coin toss winner
    ongoingGames.cell(row = rowNum, column = 27).value = "NONE"  # coin toss decision
    ongoingGames.cell(row = rowNum, column = 28).value = 1  # quarter
    ongoingGames.cell(row = rowNum, column = 29).value = "7:00"  # time
    ongoingGames.cell(row = rowNum, column = 30).value = homeTeamInfo["name"] + " 35"  # yard line
    ongoingGames.cell(row = rowNum, column = 31).value = 1  # down
    ongoingGames.cell(row = rowNum, column = 32).value = 10  # distance
    ongoingGames.cell(row = rowNum, column = 33).value = homeTeamInfo["name"]  # possession
    ongoingGames.cell(row = rowNum, column = 34).value = homeTeamInfo["user"]  # waiting on
    ongoingGames.cell(row = rowNum, column = 35).value = "KICKOFF"  # play type
    ongoingGames.cell(row = rowNum, column = 36).value = 0  # offensive number
    ongoingGames.cell(row = rowNum, column = 37).value = 0  # defensive number
    ongoingGames.cell(row = rowNum, column = 38).value = "PLAYING"  # game status
    ongoingGames.cell(row = rowNum, column = 39).value = "YES"  # clock stopped
    ongoingGames.cell(row = rowNum, column = 40).value = "NONE"  # gist link
    ongoingGames.cell(row = rowNum, column = 41).value = "NONE"  # embedded message
    ongoingGames.cell(row = rowNum, column = 42).value = "YES"  # number submitted
    ongoingGames.cell(row = rowNum, column = 43).value = "NO"  # halftime
    openpyxlGameWorkbook.save('game_database.xlsx')
    
def getGameInfoDM(user):
    """
    Get the game info from the database based on the user's discord name
    
    """
    
    rowNum = 1
    found = False
    for cell in ongoingGames['D']:
        if cell.value == str(user):
            found = True
            break
        else:
            rowNum = rowNum + 1
    if found is False:
        rowNum = 1
        for cell in ongoingGames['P']:
            if cell.value == str(user):
                break
            else:
                rowNum = rowNum + 1
    
    gameInfo = {"channel id": ongoingGames.cell(row = rowNum, column = 1).value, 
               "home name": ongoingGames.cell(row = rowNum, column = 2).value,
               "home nickname": ongoingGames.cell(row = rowNum, column = 3).value,
               "home user": ongoingGames.cell(row = rowNum, column = 4).value,
               "home offensive playbook": ongoingGames.cell(row = rowNum, column = 5).value,
               "home defensive playbook": ongoingGames.cell(row = rowNum, column = 6).value,
               "home score": ongoingGames.cell(row = rowNum, column = 7).value,
               "home timeouts": ongoingGames.cell(row = rowNum, column = 8).value,
               "home total yards": ongoingGames.cell(row = rowNum, column = 9).value,
               "home passing yards": ongoingGames.cell(row = rowNum, column = 10).value,
               "home rushing yards": ongoingGames.cell(row = rowNum, column = 11).value,
               "home interceptions": ongoingGames.cell(row = rowNum, column = 12).value,
               "home fumbles": ongoingGames.cell(row = rowNum, column = 13).value,
               "away name": ongoingGames.cell(row = rowNum, column = 14).value,
               "away nickname": ongoingGames.cell(row = rowNum, column = 15).value,
               "away user": ongoingGames.cell(row = rowNum, column = 16).value,
               "away offensive playbook": ongoingGames.cell(row = rowNum, column = 17).value,
               "away defensive playbook": ongoingGames.cell(row = rowNum, column = 18).value,
               "away score": ongoingGames.cell(row = rowNum, column = 19).value,
               "away timeouts": ongoingGames.cell(row = rowNum, column = 20).value,
               "away total yards": ongoingGames.cell(row = rowNum, column = 21).value,
               "away passing yards": ongoingGames.cell(row = rowNum, column = 22).value,
               "away rushing yards": ongoingGames.cell(row = rowNum, column = 23).value,
               "away interceptions": ongoingGames.cell(row = rowNum, column = 24).value,
               "away fumbles": ongoingGames.cell(row = rowNum, column = 25).value,
               "coin toss winner": ongoingGames.cell(row = rowNum, column = 26).value,
               "coin toss decision": ongoingGames.cell(row = rowNum, column = 27).value,
               "quarter": ongoingGames.cell(row = rowNum, column = 28).value,
               "time": ongoingGames.cell(row = rowNum, column = 29).value,
               "yard line": ongoingGames.cell(row = rowNum, column = 30).value,
               "down": ongoingGames.cell(row = rowNum, column = 31).value,
               "distance": ongoingGames.cell(row = rowNum, column = 32).value,
               "possession": ongoingGames.cell(row = rowNum, column = 33).value,
               "waiting on": ongoingGames.cell(row = rowNum, column = 34).value,
               "play type": ongoingGames.cell(row = rowNum, column = 35).value,
               "offensive number": ongoingGames.cell(row = rowNum, column = 36).value,
               "defensive number": ongoingGames.cell(row = rowNum, column = 37).value,
               "game status": ongoingGames.cell(row = rowNum, column = 38).value,
               "clock stopped": ongoingGames.cell(row = rowNum, column = 39).value,
               "gist link": ongoingGames.cell(row = rowNum, column = 40).value,
               "embedded message": ongoingGames.cell(row = rowNum, column = 41).value,
               "number submitted": ongoingGames.cell(row = rowNum, column = 42).value,
               "halftime": ongoingGames.cell(row = rowNum, column = 43).value}
    
    return gameInfo


def getGameInfo(channel):
    """
    Get the game info from the database

    """

    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1

    rowNum = rowNum + 1

    gameInfo = {"channel id": ongoingGames.cell(row=rowNum, column=1).value,
                "home name": ongoingGames.cell(row=rowNum, column=2).value,
                "home nickname": ongoingGames.cell(row=rowNum, column=3).value,
                "home user": ongoingGames.cell(row=rowNum, column=4).value,
                "home offensive playbook": ongoingGames.cell(row=rowNum, column=5).value,
                "home defensive playbook": ongoingGames.cell(row=rowNum, column=6).value,
                "home score": ongoingGames.cell(row=rowNum, column=7).value,
                "home timeouts": ongoingGames.cell(row=rowNum, column=8).value,
                "home total yards": ongoingGames.cell(row=rowNum, column=9).value,
                "home passing yards": ongoingGames.cell(row=rowNum, column=10).value,
                "home rushing yards": ongoingGames.cell(row=rowNum, column=11).value,
                "home interceptions": ongoingGames.cell(row=rowNum, column=12).value,
                "home fumbles": ongoingGames.cell(row=rowNum, column=13).value,
                "away name": ongoingGames.cell(row=rowNum, column=14).value,
                "away nickname": ongoingGames.cell(row=rowNum, column=15).value,
                "away user": ongoingGames.cell(row=rowNum, column=16).value,
                "away offensive playbook": ongoingGames.cell(row=rowNum, column=17).value,
                "away defensive playbook": ongoingGames.cell(row=rowNum, column=18).value,
                "away score": ongoingGames.cell(row=rowNum, column=19).value,
                "away timeouts": ongoingGames.cell(row=rowNum, column=20).value,
                "away total yards": ongoingGames.cell(row=rowNum, column=21).value,
                "away passing yards": ongoingGames.cell(row=rowNum, column=22).value,
                "away rushing yards": ongoingGames.cell(row=rowNum, column=23).value,
                "away interceptions": ongoingGames.cell(row=rowNum, column=24).value,
                "away fumbles": ongoingGames.cell(row=rowNum, column=25).value,
                "coin toss winner": ongoingGames.cell(row=rowNum, column=26).value,
                "coin toss decision": ongoingGames.cell(row=rowNum, column=27).value,
                "quarter": ongoingGames.cell(row=rowNum, column=28).value,
                "time": ongoingGames.cell(row=rowNum, column=29).value,
                "yard line": ongoingGames.cell(row=rowNum, column=30).value,
                "down": ongoingGames.cell(row=rowNum, column=31).value,
                "distance": ongoingGames.cell(row=rowNum, column=32).value,
                "possession": ongoingGames.cell(row=rowNum, column=33).value,
                "waiting on": ongoingGames.cell(row=rowNum, column=34).value,
                "play type": ongoingGames.cell(row=rowNum, column=35).value,
                "offensive number": ongoingGames.cell(row=rowNum, column=36).value,
                "defensive number": ongoingGames.cell(row=rowNum, column=37).value,
                "game status": ongoingGames.cell(row=rowNum, column=38).value,
                "clock stopped": ongoingGames.cell(row=rowNum, column=39).value,
                "gist link": ongoingGames.cell(row=rowNum, column=40).value,
                "embedded message": ongoingGames.cell(row=rowNum, column=41).value,
                "number submitted": ongoingGames.cell(row=rowNum, column=42).value,
                "halftime": ongoingGames.cell(row=rowNum, column=43).value}

    return gameInfo

def getGameInfoTeam(homeTeam):
    """
    Get the game info from the database based on the home team
    
    """
    
    rowNum = 1
    for cell in ongoingGames['B']:
        if cell.value == str(homeTeam):
            break
        else:
            rowNum = rowNum + 1
    
    gameInfo = {"channel id": ongoingGames.cell(row = rowNum, column = 1).value, 
               "home name": ongoingGames.cell(row = rowNum, column = 2).value,
               "home nickname": ongoingGames.cell(row = rowNum, column = 3).value,
               "home user": ongoingGames.cell(row = rowNum, column = 4).value,
               "home offensive playbook": ongoingGames.cell(row = rowNum, column = 5).value,
               "home defensive playbook": ongoingGames.cell(row = rowNum, column = 6).value,
               "home score": ongoingGames.cell(row = rowNum, column = 7).value,
               "home timeouts": ongoingGames.cell(row = rowNum, column = 8).value,
               "home total yards": ongoingGames.cell(row = rowNum, column = 9).value,
               "home passing yards": ongoingGames.cell(row = rowNum, column = 10).value,
               "home rushing yards": ongoingGames.cell(row = rowNum, column = 11).value,
               "home interceptions": ongoingGames.cell(row = rowNum, column = 12).value,
               "home fumbles": ongoingGames.cell(row = rowNum, column = 13).value,
               "away name": ongoingGames.cell(row = rowNum, column = 14).value,
               "away nickname": ongoingGames.cell(row = rowNum, column = 15).value,
               "away user": ongoingGames.cell(row = rowNum, column = 16).value,
               "away offensive playbook": ongoingGames.cell(row = rowNum, column = 17).value,
               "away defensive playbook": ongoingGames.cell(row = rowNum, column = 18).value,
               "away score": ongoingGames.cell(row = rowNum, column = 19).value,
               "away timeouts": ongoingGames.cell(row = rowNum, column = 20).value,
               "away total yards": ongoingGames.cell(row = rowNum, column = 21).value,
               "away passing yards": ongoingGames.cell(row = rowNum, column = 22).value,
               "away rushing yards": ongoingGames.cell(row = rowNum, column = 23).value,
               "away interceptions": ongoingGames.cell(row = rowNum, column = 24).value,
               "away fumbles": ongoingGames.cell(row = rowNum, column = 25).value,
               "coin toss winner": ongoingGames.cell(row = rowNum, column = 26).value,
               "coin toss decision": ongoingGames.cell(row = rowNum, column = 27).value,
               "quarter": ongoingGames.cell(row = rowNum, column = 28).value,
               "time": ongoingGames.cell(row = rowNum, column = 29).value,
               "yard line": ongoingGames.cell(row = rowNum, column = 30).value,
               "down": ongoingGames.cell(row = rowNum, column = 31).value,
               "distance": ongoingGames.cell(row = rowNum, column = 32).value,
               "possession": ongoingGames.cell(row = rowNum, column = 33).value,
               "waiting on": ongoingGames.cell(row = rowNum, column = 34).value,
               "play type": ongoingGames.cell(row = rowNum, column = 35).value,
               "offensive number": ongoingGames.cell(row = rowNum, column = 36).value,
               "defensive number": ongoingGames.cell(row = rowNum, column = 37).value,
               "game status": ongoingGames.cell(row = rowNum, column = 38).value,
               "clock stopped": ongoingGames.cell(row = rowNum, column = 39).value,
               "gist link": ongoingGames.cell(row = rowNum, column = 40).value,
               "embedded message": ongoingGames.cell(row = rowNum, column = 41).value,
               "number submitted": ongoingGames.cell(row = rowNum, column = 42).value,
               "halftime": ongoingGames.cell(row = rowNum, column = 43).value}
    
    return gameInfo

def updateHomeScore(channel, score):
    """
    Update the home score
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 7).value = score  # home score
    openpyxlGameWorkbook.save('game_database.xlsx')

def updateHomeTimeouts(channel, timeouts):
    """
    Update the home timeouts
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 8).value = timeouts  # home timeouts
    openpyxlGameWorkbook.save('game_database.xlsx')
    
def updateAwayScore(channel, score):
    """
    Update the away score
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 19).value = score  # away score
    openpyxlGameWorkbook.save('game_database.xlsx')
    
def updateAwayTimeouts(channel, timeouts):
    """
    Update the away timeouts
    
    """

    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 20).value = timeouts  # away timeouts
    openpyxlGameWorkbook.save('game_database.xlsx')
    
def updateCoinTossWinner(channel, winner):
    """
    Update the coin toss winner
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 26).value = winner  # coin toss winner
    openpyxlGameWorkbook.save('game_database.xlsx')
    
def updateCoinTossDecision(channel, decision):
    """
    Update the coin toss decision
    
    """

    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 27).value = decision  # coin toss decision
    openpyxlGameWorkbook.save('game_database.xlsx')
    

def updateQuarter(channel, quarter):
    """
    Update the quarter
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 28).value = quarter  # coin toss decision
    openpyxlGameWorkbook.save('game_database.xlsx')
    

def updateTime(channel, time):
    """
    Update the time
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 29).value = time  # coin toss decision
    openpyxlGameWorkbook.save('game_database.xlsx')
  

def updateNormalKickoffBallLocation(channel, homeTeam, awayTeam, result, possession):
    """
    Update the normal kickoff location
    
    """

    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    yardLine = ""
    
    if possession == homeTeam and result != "50" and result != "Returned TD" and result != "Fumble" and result != "Touchdown":
        yardLine = awayTeam + " " + result
    elif possession == awayTeam and result != "50" and result != "Returned TD" and result != "Fumble" and result != "Touchdown":
        yardLine = homeTeam + " " + result
    elif result == "50":
        yardLine = result
    elif possession == homeTeam and result == "Returned TD":
        yardLine = homeTeam + " 3"
    elif possession == awayTeam and result == "Returned TD":
        yardLine = awayTeam + " 3"
    elif possession == homeTeam and result == "Fumble":
        yardLine = awayTeam + " 20"
    elif possession == awayTeam and result == "Fumble":
        yardLine = homeTeam + " 20"
    elif possession == homeTeam and result == "Touchdown":
        yardLine = awayTeam + " 3"
    elif possession == awayTeam and result == "Touchdown":
        yardLine = homeTeam + " 3"
    
    ongoingGames.cell(row = rowNum, column = 30).value = yardLine  # yard line
    openpyxlGameWorkbook.save('game_database.xlsx')


def updateSquibKickoffBallLocation(channel, homeTeam, awayTeam, result, possession):
    """
    Update the squib kickoff ball location
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    yardLine = ""
    
    if possession == homeTeam and result != "50" and result != "Returned TD" and result != "Fumble" and result != "Touchdown":
        yardLine = homeTeam + " " + result
    elif possession == awayTeam and result != "50" and result != "Returned TD" and result != "Fumble" and result != "Touchdown":
        yardLine = awayTeam + " " + result
    elif result == "50":
        yardLine = result
    elif possession == homeTeam and result == "Returned TD":
        yardLine = homeTeam + " 3"
    elif possession == awayTeam and result == "Returned TD":
        yardLine = awayTeam + " 3"
    elif possession == homeTeam and result == "Fumble":
        yardLine = awayTeam + " 30"
    elif possession == awayTeam and result == "Fumble":
        yardLine = homeTeam + " 30"
    elif possession == homeTeam and result == "Touchdown":
        yardLine = awayTeam + " 3"
    elif possession == awayTeam and result == "Touchdown":
        yardLine = homeTeam + " 3"
    
    ongoingGames.cell(row = rowNum, column = 30).value = yardLine  # yard line
    openpyxlGameWorkbook.save('game_database.xlsx')
    

def updateOnsideKickoffBallLocation(channel, homeTeam, awayTeam, result, possession):
    """
    Update the onside kickoff ball location
    
    """
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    yardLine = ""

    if possession == homeTeam and result == "Recovered":
        yardLine = homeTeam + " 45"
        updatePossession(channel, homeTeam)
    elif possession == homeTeam and result == "No Good":
        yardLine = homeTeam + " 45"
        updatePossession(channel, awayTeam)
    elif possession == awayTeam and result == "Recovered":
        yardLine = awayTeam + " 45"
        updatePossession(channel, awayTeam)
    elif possession == awayTeam and result == "No Good":
        yardLine = awayTeam + " 45"
        updatePossession(channel, homeTeam)
    elif possession == homeTeam and result == "Returned TD":
        yardLine = homeTeam + " 3"
        updatePossession(channel, awayTeam)
    elif possession == awayTeam and result == "Returned TD":
        yardLine = awayTeam + " 3"
        updatePossession(channel, homeTeam)
    
    ongoingGames.cell(row = rowNum, column = 30).value = yardLine  # yard line
    openpyxlGameWorkbook.save('game_database.xlsx')

    
def updateBallLocation(channel, yardLine):
    """
    Update the ball location
    
    """
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 30).value = yardLine  # yard line
    openpyxlGameWorkbook.save('game_database.xlsx')


def updateDown(channel, down):
    """
    Update the down
    
    """   
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 31).value = down  # down
    openpyxlGameWorkbook.save('game_database.xlsx')
    
   
def updateDistance(channel, distance):
    """
    Update the distance
    
    """ 
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 32).value = distance  # distance
    openpyxlGameWorkbook.save('game_database.xlsx') 
    
     
def updatePossession(channel, possessingTeam):
    """
    Update the possession
    
    """   
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 33).value = possessingTeam  # possession
    openpyxlGameWorkbook.save('game_database.xlsx')


def updateWaitingOnUser(user):
    """
    Update the team waiting on

    """

    rowNum = 1
    foundGame = False
    for cell in ongoingGames['D']:
        if cell.value == str(user):
            foundGame = True
            break
        else:
            rowNum = rowNum + 1
    if foundGame is False:
        rowNum = 1
        for cell in ongoingGames['P']:
            if cell.value == str(user):
                break
            else:
                rowNum = rowNum + 1

    waitingOn = user

    ongoingGames.cell(row=rowNum, column=34).value = waitingOn  # waiting on
    openpyxlGameWorkbook.save('game_database.xlsx')
 
def updateWaitingOn(channel):
    """
    Update the team waiting on
    
    """   
    
    rowNum = 1
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
    
    waitingOn = "NO ONE"
    
    gameInfo = getGameInfo(channel)
    if gameInfo["play type"] == "KICKOFF":
        if gameInfo["possession"] == gameInfo["home name"]:
            waitingOn = gameInfo["away user"]
        elif gameInfo["possession"] == gameInfo["away name"]:
            waitingOn = gameInfo["home user"]
    else:
        if gameInfo["possession"] == gameInfo["home name"]:
            waitingOn = gameInfo["home user"]
        elif gameInfo["possession"] == gameInfo["away name"]:
            waitingOn = gameInfo["away user"]
    
    ongoingGames.cell(row = rowNum, column = 34).value = waitingOn  # waiting on
    openpyxlGameWorkbook.save('game_database.xlsx')
    
    
def updateWaitingOnKickoffDM(channel):
    """
    Update the team waiting on for when the kickoff number is DMed
    
    """   
    
    rowNum = 1
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
    
    waitingOn = "NO ONE"
    
    gameInfo = getGameInfo(channel)
    if gameInfo["possession"] == gameInfo["home name"]:
        waitingOn = gameInfo["home user"]
    elif gameInfo["possession"] == gameInfo["away name"]:
        waitingOn = gameInfo["away user"]
    
    ongoingGames.cell(row = rowNum, column = 34).value = waitingOn  # waiting on
    openpyxlGameWorkbook.save('game_database.xlsx')
    
    
def updatePlayType(channel, playType):
    """
    Update the play type
    
    """  
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 35).value = playType  # play type
    openpyxlGameWorkbook.save('game_database.xlsx')
      
def updateOffensiveNumber(channel, number):
    """
    Update the offensive number
    
    """  
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 36).value = number  # offensive number
    openpyxlGameWorkbook.save('game_database.xlsx')
    
   
def updateDefensiveNumber(channel, number):
    """
    Update the defensive number
    
    """ 
    
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 37).value = number  # defensive number
    openpyxlGameWorkbook.save('game_database.xlsx')
    
    
def updateGameStatus(channel, status):
    """
    Update the game status flag
    
    """
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 38).value = status  # possession
    openpyxlGameWorkbook.save('game_database.xlsx')
    
   
def updateClockStopped(channel, status):
    """
    Update the game clock stopped flag
    
    """ 
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 39).value = status  # possession
    openpyxlGameWorkbook.save('game_database.xlsx')
  

def updateGist(channel, gist):
    """
    Update the GitHub gist link
    
    """ 
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 40).value = str(gist)  # gist
    openpyxlGameWorkbook.save('game_database.xlsx')
    
    
def updateEmbeddedMessage(channel, embedded):
    """
    Update the Discord embedded message link
    
    """ 
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 41).value = str(embedded)  # embedded
    openpyxlGameWorkbook.save('game_database.xlsx')
    

def updateNumberSubmitted(channel, status):
    """
    Update if the defensive number has been submitted
    
    """ 
    rowNum = 0
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 42).value = str(status)  # status
    openpyxlGameWorkbook.save('game_database.xlsx')
  
    
def updateHalftime(channel, status):
    """
    Update if the defensive number has been submitted
    
    """ 
    rowNum = 1
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        else:
            rowNum = rowNum + 1
    
    ongoingGames.cell(row = rowNum, column = 43).value = str(status)  # status
    openpyxlGameWorkbook.save('game_database.xlsx')
    

def copyGameData(channel): 
    """
    Copy the game data to a list
    
    """
    
    rowNum = 0
    noGame = False
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        if cell.value is None or cell.value == "":
            noGame = True
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    if not noGame:
        data = []
        
        for i in range(1, 43):
            cellValue = ongoingGames.cell(row = rowNum, column = i).value
            if cellValue is None or cellValue == "" or cellValue == " ":
                break
            else:
                data.append(cellValue)
                
        return data
    else:
        return "NO GAME FOUND"

def deleteGameData(channel):
    """
    Delete the game data from a list
    
    """
    rowNum = 0
    noGame = False
    for cell in ongoingGames['A']:
        if cell.value == str(channel.id):
            break
        if cell.value is None or cell.value == "":
            noGame = True
            break
        else:
            rowNum = rowNum + 1
    
    rowNum = rowNum + 1
    
    if not noGame:
        ongoingGames.delete_rows(rowNum)
        openpyxlGameWorkbook.save('game_database.xlsx')
    
        
def pasteGameData(data):
    """
    Paste the game data in a new location
    
    """
    
    rowNum = 0
    for cell in finishedGames['A']:
        if cell.value is None or cell.value == "":
            break
        else:
            rowNum = rowNum + 1
            
    rowNum = rowNum + 1
    
    for i in range(1, len(data)):
        finishedGames.cell(row = rowNum, column = i).value = data[i-1]
        
    openpyxlGameWorkbook.save('game_database.xlsx')

   
def checkUserFree(user):
    """
    Iterate through the ongoing games and verify the user does not have a game going
    
    """ 

    for cell in ongoingGames['D']:
        if cell.value == user:
            return False
        elif cell.value is None or cell.value == "":
            break
    
    for cell in ongoingGames['P']:
        if cell.value == user:
            return False
        elif cell.value is None or cell.value == "":
            break
            
    return True
